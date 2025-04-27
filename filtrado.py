# -*- coding: utf-8 -*-
"""
Created on Sat Apr 12 21:04:37 2025

@author: Jebus
"""
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font

import pandas as pd
from datetime import datetime

def parse_wos_entries(file_path):
    entries = []
    current_entry = {}
    current_key = None

    with open(file_path, 'r', encoding='utf-8-sig') as f:
        line = f.readline()
        if "FN Clarivate" not in line:
            print("Fuente no soportada")
            return("Fuente no soportada")
        for line in f:
            line = line.rstrip()
            if not line:
                continue

            if len(line) >= 3 and line[:2].isalpha() and line[2] == ' ':
                key = line[:2]
                value = line[3:].strip()

                if key == "PT" and current_entry:
                    # Guardamos la entrada anterior y empezamos una nueva
                    entries.append(current_entry)
                    current_entry = {}

                current_key = key

                if key in current_entry:
                    # Si ya existe, lo convertimos en lista
                    if isinstance(current_entry[key], list):
                        current_entry[key].append(value)
                    else:
                        current_entry[key] = [current_entry[key], value]
                else:
                    current_entry[key] = value
            else:
                # Continuación de la línea anterior
                if current_key:
                    continuation = line.strip()
                    if isinstance(current_entry[current_key], list):
                        current_entry[current_key].append(continuation)
                    else:
                        current_entry[current_key] = [current_entry[current_key], continuation]

        # Agregamos la última entrada
        if current_entry:
            entries.append(current_entry)

    return entries


def buscar_entradas(entries, modo='OR', **criterios):
    """
    Busca entradas en la lista de diccionarios según los criterios dados.

    Parámetros:
        entries (list): Lista de diccionarios (entradas).
        modo (str): 'AND' (todos los criterios) o 'OR' (al menos uno).
        criterios (dict): Claves y valores a buscar. Ej:
                          TI="food processing", AU="Barba", PY="2021"

    Retorna:
        Lista de entradas que coinciden con los criterios según el modo.
    """
    if modo not in ('AND', 'OR'):
        raise ValueError("El parámetro 'modo' debe ser 'AND' o 'OR'.")

    resultados = []

    for entrada in entries:
        condiciones = []

        for campo, valor_buscado in criterios.items():
            if campo not in entrada:
                condiciones.append(False)
                continue

            campo_valor = entrada[campo]
            if isinstance(campo_valor, list):
                campo_valor = " ".join(map(str, campo_valor))

            coincidencia = valor_buscado.lower() in campo_valor.lower()
            condiciones.append(coincidencia)

        # Combinar las condiciones según el modo
        if modo == 'AND' and all(condiciones):
            resultados.append(entrada)
        elif modo == 'OR' and any(condiciones):
            resultados.append(entrada)

    return resultados


def exportar_a_excel(resultados, filtro, archivo_salida='resultados.xlsx'):
    if not resultados:
        print("No hay resultados para exportar.")
        return ("No hay resultados para exportar.")

    # Generar un nuevo archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.title = filtro
    
    # Definir tipo de letra para encabezados y contenido
    fuente_encabezado = Font(name="System", bold=True, size=12)
    fuente_contenido = Font(name="System", size=11)

    # Obtener todas las claves únicas de los diccionarios
    claves = set()
    for entrada in resultados:
        claves.update(entrada.keys())

    claves = sorted(claves)

    # Escribir los encabezados
    for col, clave in enumerate(claves, 1):        
        celda = ws.cell(row=1, column=col, value=clave)
        celda.font = fuente_encabezado

    # Escribir los valores
    for fila, entrada in enumerate(resultados, start=2):
        for col, clave in enumerate(claves, 1):
            valor = entrada.get(clave, "")
            if isinstance(valor, list):
                valor = "; ".join(valor)
            celda = ws.cell(row=fila, column=col, value=valor)
            celda.font = fuente_contenido

    # Ajustar ancho de columnas automáticamente
    for i, clave in enumerate(claves, 1):
        longitud = max(len(str(clave)), 15)
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = longitud

    wb.save(archivo_salida)
    print(f"Archivo Excel guardado como: {archivo_salida}")

def parse_sco_entries(file_path):
    file_tmp = pd.read_csv(file_path, index_col = False)
    file_tmp.columns = [col.strip().replace(" ", "_") for col in file_tmp.columns]    
    original = ['Author_Keywords', 'Index_Keywords']
    header = list(file_tmp.columns.values)
    if original[0] in header or original[1] in header:
        return file_tmp, header
    else:
        print("Fuente no soportada")
        return("Fuente no soportada")
    

def buscar_en_csv(df, modo='OR', **criterios):
    """
    Busca filas que cumplan los criterios dados, combinados por AND o OR.
    
    Para datos de SCOPUS

    Parámetros:
        df (DataFrame): El DataFrame original.
        modo (str): 'AND' (por defecto) o 'OR', para combinar los filtros.
        criterios (dict): Diccionario con columnas como claves y valores a buscar.

    Retorna:
        DataFrame filtrado o un mensaje si no hay resultados.
    """
    if not criterios:
        print("⚠️ No se especificaron criterios.")
        return df

    if modo not in ('AND', 'OR'):
        raise ValueError("El modo debe ser 'AND' o 'OR'.")

    condiciones = []

    for columna, valor_buscado in criterios.items():
        if columna not in df.columns:
            print(f"⚠️ La columna '{columna}' no existe en el DataFrame.")
            continue

        # Generar la condición para esta columna
        condicion = df[columna].astype(str).str.contains(str(valor_buscado), case=False, na=False)
        condiciones.append(condicion)

    if not condiciones:
        return "No hay resultados para exportar."

    # Combinar condiciones con AND o OR
    if modo == 'AND':
        combinada = condiciones[0]
        for c in condiciones[1:]:
            combinada &= c
    else:  # modo == 'OR'
        combinada = condiciones[0]
        for c in condiciones[1:]:
            combinada |= c

    resultado = df[combinada]

    if resultado.empty:
        return "No hay resultados para exportar."
    
    return resultado


def a_excel_fuente(resultados, filtro, fecha):
    #    SCOPUS
    archivo = f"resultados_{filtro}_{fecha}.xlsx"
    
    # Guardar con pandas
    resultados.to_excel(archivo, index=False)
    
    # Cargar con openpyxl para aplicar formato
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    ws.title = filtro
    
    # Estilo de fuente
    fuente = Font(name="System", size=11)
    
    # Aplicar estilo a todas las celdas con datos
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for celda in row:
            celda.font = fuente
    
    # Guardar el archivo modificado
    wb.save(archivo)

def export_to_wos_format(records, output_file):
    """
    Exporta una lista de diccionarios a un archivo de texto plano con formato Web of Science.

    Args:
        records: lista de diccionarios, cada uno representando un artículo.
        output_file: nombre del archivo de salida (.txt).
    """
    with open(output_file, 'w', encoding='utf-8') as f:
        for record in records:
            # Tipo de publicación (por defecto 'J' de Journal Article)
            pt_value = record.get('PT', 'J')
            f.write(f"PT {pt_value}\n")
            
            # Escribir campos
            if 'AU' in record and record['AU']:
                autores = record['AU']
                if isinstance(autores, list):
                    for autor in autores:
                        f.write(f"AU {autor}\n")
                else:
                    f.write(f"AU {autores}\n")
                    
            if 'TI' in record and record['TI']:
                f.write(f"TI {record['TI']}\n")
                
            if 'SO' in record and record['SO']:
                f.write(f"SO {record['SO']}\n")
                
            if 'DE' in record and record['DE']:
                keywords = record['DE']
                if isinstance(keywords, list):
                    for keyword in keywords:
                        f.write(f"DE {keyword}\n")
                else:
                    f.write(f"DE {keywords}\n")
                    
            if 'AB' in record and record['AB']:
                f.write(f"AB {record['AB']}\n")
                
            if 'PY' in record and record['PY']:
                f.write(f"PY {record['PY']}\n")
                
            if 'VL' in record and record['VL']:
                f.write(f"VL {record['VL']}\n")
                
            if 'IS' in record and record['IS']:
                f.write(f"IS {record['IS']}\n")
                
            if 'BP' in record and record['BP']:
                f.write(f"BP {record['BP']}\n")
                
            if 'EP' in record and record['EP']:
                f.write(f"EP {record['EP']}\n")
                
            if 'DI' in record and record['DI']:
                f.write(f"DI {record['DI']}\n")
            
            # Terminar registro
            f.write('ER\n\n')
        
        # Terminar archivo
        f.write('EF\n')

    

def procesar_wos(file_path, filtro, modo = 'OR'): 
    fecha = datetime.now().strftime("%d%m%y%H%M")
    listado = parse_wos_entries(file_path)   
    
    datos = {}
    if 'A' in filtro:
        datos['DE'] = filtro.get('A')
    if 'B' in filtro:
        datos['ID'] = filtro.get('B')
    if 'C' in filtro:
        datos['AB'] = filtro.get('C')
    
    filtro = next(iter(filtro.values()))
    resultados = buscar_entradas(listado, modo='OR', **datos)   
    export_to_wos_format(resultados, f"resultados_{filtro}_{fecha}.txt")     
    resultados = exportar_a_excel(resultados, filtro, f"resultados_{filtro}_{fecha}.xlsx")
    if isinstance(resultados, str):
        return(resultados)
    
    """
    if modo == 'OR':
        resultados = buscar_entradas(listado, modo='OR', DE = filtro, ID = filtro)        
        resultados = exportar_a_excel(resultados, filtro, f"resultados_{filtro}_{fecha}.xlsx")
        if isinstance(resultados, str):
            return(resultados)
    
    if modo == '1':
        resultados = buscar_entradas(listado, modo='OR', DE = filtro)        
        resultados = exportar_a_excel(resultados, filtro, f"resultados_{filtro}_{fecha}.xlsx")
        if isinstance(resultados, str):
            return(resultados)
    
    if modo == '2':
        resultados = buscar_entradas(listado, modo='OR', ID = filtro)        
        resultados = exportar_a_excel(resultados, filtro, f"resultados_{filtro}_{fecha}.xlsx")
        if isinstance(resultados, str):
            return(resultados)
    
    if modo == '3':
        resultados = buscar_entradas(listado, modo='OR', AB = filtro)        
        resultados = exportar_a_excel(resultados, filtro, f"resultados_{filtro}_{fecha}.xlsx")
        if isinstance(resultados, str):
            return(resultados)
    """
    return (f"resultados_{filtro}_{fecha}.xlsx")

def procesar_sco(file_path, filtro, modo='OR'):  
    fecha = datetime.now().strftime("%d%m%y%H%M")
    file_tmp, header = parse_sco_entries(file_path)          
    
    datos = {}
    if 'A' in filtro:
        datos['Author_Keywords'] = filtro.get('A')
    if 'B' in filtro:
        datos['Index_Keywords'] = filtro.get('B')
    if 'C' in filtro:
        datos['Abstract'] = filtro.get('C')

    
    resultados = buscar_en_csv(file_tmp, modo='OR', **datos)
    if isinstance(resultados, str):
        return(resultados)
    
    """
    if modo == 'OR':
        resultados = buscar_en_csv(file_tmp, modo ='OR', 
                                   Author_Keywords = filtro, 
                                   Index_Keywords = filtro,
                                   Abstract = filtro)
        if isinstance(resultados, str):
            return(resultados)
    
    if modo == '1':
        resultados = buscar_en_csv(file_tmp, modo='OR', Author_Keywords=filtro)
        if isinstance(resultados, str):
            return(resultados)
    if modo == '2':
        resultados = buscar_en_csv(file_tmp, modo='OR', Index_Keywords=filtro)
        if isinstance(resultados, str):
            return(resultados)
    
    if modo == '3':
        resultados = buscar_en_csv(file_tmp, modo='OR', Abstract=filtro)
        if isinstance(resultados, str):
            return(resultados)
    """
    
    for col in header:
        if col not in resultados.columns:
            resultados[col] = ""
    
    # Reordenar
    df_result = resultados[header]
    df_result.to_csv(f"resultados_{filtro}_{fecha}.csv", index=False, encoding='utf-8-sig')
    
    filtro = next(iter(filtro.values()))
    a_excel_fuente(resultados, filtro, fecha)
    return (f"resultados_{filtro}_{fecha}.xlsx")
    


def debug():
    
    wos = 1
    debug = 1
    
    if wos:       
        T0 = 'savedrecs_2Articulos.txt'
        T1 = 'WOS_Design_070425.txt'
        T2 = 'WOS_Mechatronic design_070425.txt'
        T3 = "doc.txt"
        T4 = 'scopus_IA_Math_Edu_120425.csv'
        
        if debug:
            file_path = T0
            listado = parse_wos_entries(file_path)
        else:
            file_path = input("Esciba el nombre del archivo a analizar.\n") + ".txt"
            listado = parse_wos_entries(file_path)
            filtro = input("Ingrese la palabra clave.\n")
            resultados = buscar_entradas(listado, modo='AND', DE = filtro, ID = filtro)
            exportar_a_excel(resultados, "resultados_filtrados_and.xlsx")
            resultados = buscar_entradas(listado, modo='OR', DE = filtro, ID = filtro)
            exportar_a_excel(resultados, "resultados_filtrados_or.xlsx")
    
    else:
          if debug:
              file_path = 'scopus_IA_Math_Edu_120425.csv'
              file_tmp, header = parse_sco_entries(file_path)
              resultados_and = buscar_en_csv(file_tmp, modo='AND', Author_Keywords="industry", Index_Keywords='industry')
              resultados_or = buscar_en_csv(file_tmp, modo='OR', Author_Keywords="industry", Index_Keywords='industry')
              resultados_or.to_excel("resultados_or.xlsx", index=False)
              resultados_and.to_excel("resultados_and.xlsx", index=False)
          else:    
              pass